import openpyxl
import re

filePath = 'C:\\Users\\wk\\Desktop\\龙票易信短信模版.xlsx'
wb = openpyxl.load_workbook(filePath)
# 获取所有的工作表名
sheetNames = wb.sheetnames
print("获取所有的工作表名:", sheetNames)
smsSheet = wb[sheetNames[0]]
print("获取指定的工作表:", smsSheet)
maxRow = smsSheet.max_row
print("最大行数:", maxRow)
maxColumn = smsSheet.max_column
print("最大列数:", maxColumn)
print()

regEx = "\\{(.*?)\\}"


def replaceStr(content):
    resp = re.findall(r"\{(.*?)\}", content)
    # print(resp)
    if len(resp) > 0:
        for i in range(len(resp)):
            content = content.replace(resp[i], str(i), 1)
            pass
        pass
    pass
    # print(content)
    return content


# 循环工作簿的所有行
for row in smsSheet.iter_rows():
    bColumn = row[1]  # 编号所在的列
    code = bColumn.value
    if code != '编号':  # 排除第一行
        gColumn = row[6]  # 发送内容所在的列
        content = gColumn.value
        # 打印
        title = row[4].value
        if content is not None:
            code = bColumn.value.replace("-", "")
            h = row[7].value
            i = row[8].value
            if h is None:
                h = '系统'
            if i is None:
                i = ''
            receiver = h + '-' + i

            f = row[5].value
            if f is None:
                f = ''

            # print(type(code),type(title),type(content))
            if title is None:
                title = ""
            sendTime = title + "-" + f
            memo = replaceStr(content)
            result = code + '("' + sendTime + '",' + '"' + memo + '","' + content + '","' + receiver + '"),'
            # print(code, content)
            # print(result)
            print(result)
    pass
