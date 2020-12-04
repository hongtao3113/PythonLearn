#
from openpyxl import load_workbook

filePath = 'C:\\Users\\wk\\Desktop\\消息发送模板v2.xlsx'
excel = load_workbook(filePath)
print("type:", type(excel))
names = excel.sheetnames
print("names", names)
table1 = excel['短信']
'''
DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
table = excel['短信'] 正常
get_sheet_by_name('短信') 报上述错误
'''
rows = table1.max_row
cols = table1.max_column
print("rows:", rows)
print("cols:", cols)

print(table1.title)
print(table1['A1'])
print(table1['A1'].value)
