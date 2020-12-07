import openpyxl

filePath = 'C:\\Users\\wk\\Desktop\\消息发送模板v2.xlsx'
wb = openpyxl.load_workbook(filePath)
# 获取所有的工作表名
sheetNames = wb.sheetnames
print("获取所有的工作表名:", sheetNames)
smsSheet = wb[sheetNames[0]]
print("获取指定的工作表:", smsSheet)
maxRow = smsSheet.max_row
print("最大行数:", maxRow)
maxColumn = smsSheet.max_column
print("最大列数:", maxColumn)
print()
for one_column_data in smsSheet.iter_rows():
    if isinstance(one_column_data[0].value, int):
        columnValue = one_column_data[0].value
        string = 'SMS'
        if columnValue < 10:
            string = string + '0' + str(columnValue) + '('
        else:
            string = string + str(columnValue) + '('
        for i in range(maxColumn):
            if i > 1:
                v = '"' + one_column_data[i].value + '"'
                string = string + v
                if i == 3:
                    string = string + ',' + v
                if maxColumn - 1 != i:
                    string = string + ','
                else:
                    string = string + ')'
            pass
        string = string + ',\n'
        print(string)
    pass
