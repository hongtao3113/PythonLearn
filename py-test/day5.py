import openpyxl

filePath = 'C:\\Users\\wk\\Desktop\\消息发送模板v2.xlsx'
wb = openpyxl.load_workbook(filePath)

# 获取所有的工作表名
sheetNames = wb.sheetnames
print("获取所有的工作表名:", sheetNames)

smsSheet = wb[sheetNames[0]]
print("获取指定的工作表", smsSheet)

maxRow = smsSheet.max_row
print("最大行数", maxRow)
maxColumn = smsSheet.max_column
print("最大列数", maxColumn)

currentSheet = wb.active
print("当前工作表：", currentSheet)

# 通过名字访问Cell对象, 通过value属性获取值
a1 = smsSheet['A1'].value
print("通过value属性获取值:", a1)

# 通过行和列确定数据
a11 = smsSheet.cell(row=1, column=1).value
print("通过行和列确定数据:", a11)

column = openpyxl.utils.cell.get_column_letter(1)
print("column:", column)
column_name_num = openpyxl.utils.cell.column_index_from_string('a')
print("column_name_num:", column_name_num)

print(
    "------------------------------------------------------------------------------------------------------------------------")

for one_column_data in smsSheet.iter_rows():
    for i in range(maxColumn):
        print(one_column_data[i].value)
    pass

print(
    "------------------------------------------------------------------------------------------------------------------------")
dataRowNum = maxRow - 1
print("dataRowNum:", dataRowNum)
for one_row_data in smsSheet.iter_cols():
    for i in range(maxColumn):
        print(one_row_data[i].value, end="\t")
        print()
    pass

print(
    "------------------------------------------------------------------------------------------------------------------------")
print()
for i in range(maxColumn):
    print(i)
    pass

print("row = {}, column = {}".format(maxRow, maxColumn))
